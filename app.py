import io
import openpyxl
from datetime import date, timedelta
import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from models import db, User, Category, MLSize, CategoryMLMapping, Product, StockEntry, Expense

app = Flask(__name__)
app.config['SECRET_KEY'] = 'a_very_secret_shop_key_2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///shop.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def owner_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'owner':
            flash('Access denied. Owner only.', 'danger')
            return redirect(url_for('staff_dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def init_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='owner').first():
            db.session.add(User(username='owner', password_hash=generate_password_hash('owner123'), role='owner'))
        if not User.query.filter_by(username='staff').first():
            db.session.add(User(username='staff', password_hash=generate_password_hash('staff123'), role='staff'))
        
        categories = {'Whisky': 1, 'Gin': 2, 'Vodka': 3, 'Rum': 4, 'Wine': 5, 'Beer': 6}
        for cat_name, order in categories.items():
            cat = Category.query.filter_by(name=cat_name).first()
            if not cat:
                db.session.add(Category(name=cat_name, category_order=order))
            else:
                cat.category_order = order
        
        ml_sizes = [90, 180, 275, 330, 500, 650, 750]
        for ml_val in ml_sizes:
            if not MLSize.query.filter_by(value=ml_val).first():
                db.session.add(MLSize(value=ml_val))
        
        db.session.commit()
        
        beer_cat = Category.query.filter_by(name='Beer').first()
        spirits = Category.query.filter(Category.name != 'Beer').all()
        
        spirit_mls = MLSize.query.filter(MLSize.value.in_([90, 180, 275, 750])).all()
        beer_mls = MLSize.query.filter(MLSize.value.in_([330, 500, 650])).all()
        
        for cat in spirits:
            for ml in spirit_mls:
                if not CategoryMLMapping.query.filter_by(category_id=cat.id, ml_id=ml.id).first():
                    db.session.add(CategoryMLMapping(category_id=cat.id, ml_id=ml.id))
        for ml in beer_mls:
            if not CategoryMLMapping.query.filter_by(category_id=beer_cat.id, ml_id=ml.id).first():
                db.session.add(CategoryMLMapping(category_id=beer_cat.id, ml_id=ml.id))
        
        # Cleanup routine strictly stripping bad mapping natively
        bad_ml = MLSize.query.filter_by(value=750).first()
        if bad_ml:
            bad_mapping = CategoryMLMapping.query.filter_by(category_id=beer_cat.id, ml_id=bad_ml.id).first()
            if bad_mapping:
                db.session.delete(bad_mapping)
                
        db.session.commit()

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'owner':
            return redirect(url_for('owner_dashboard'))
        else:
            return redirect(url_for('staff_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            if user.role == 'owner':
                return redirect(url_for('owner_dashboard'))
            return redirect(url_for('staff_dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def get_target_date():
    today = date.today()
    working_date = today - timedelta(days=1)
    
    last_entry = StockEntry.query.filter_by(is_completed=True).order_by(StockEntry.date.desc()).first()
    
    if not last_entry:
        return working_date, working_date
        
    target_date = last_entry.date + timedelta(days=1)
    return target_date, working_date

def ensure_entries_for(target_date):
    products = Product.query.filter_by(is_active=True).all()
    created = False
    for product in products:
        entry = StockEntry.query.filter_by(product_id=product.id, date=target_date).first()
        if not entry:
            prev = StockEntry.query.filter(StockEntry.product_id == product.id, StockEntry.date < target_date).order_by(StockEntry.date.desc()).first()
            opening = prev.closing_stock if prev else 0
            new_entry = StockEntry(
                date=target_date, product_id=product.id,
                opening_stock=opening, new_stock=0, total_available_stock=opening,
                closing_stock=opening, sold=0,
                cost_price_at_entry=product.cost_price, 
                selling_price_at_entry=product.selling_price,
                is_completed=False, entry_type='Pending'
            )
            db.session.add(new_entry)
            created = True
    if created:
        db.session.commit()

@app.route('/staff/dashboard', methods=['GET', 'POST'])
@login_required
def staff_dashboard():
    target_date, working_date = get_target_date()
    
    # If target_date <= working_date, they have pending work to resolve.
    # Otherwise, they are completely up to date.
    is_up_to_date = target_date > working_date
    
    if not is_up_to_date:
        ensure_entries_for(target_date)
        
    if request.method == 'POST' and not is_up_to_date:
        action = request.form.get('action') # 'regular', 'estimated', 'closed'
        entries = StockEntry.query.filter_by(date=target_date).all()
        
        if action == 'closed':
            # Fast-forward all entries for that day as closed
            for entry in entries:
                entry.new_stock = 0
                entry.total_available_stock = entry.opening_stock
                entry.closing_stock = entry.opening_stock
                entry.sold = 0
                entry.is_completed = True
                entry.entry_type = 'Closed'
            db.session.commit()
            flash(f'Data submitted successfully for {target_date.strftime("%Y-%m-%d")}', 'success')
            return redirect(url_for('staff_dashboard'))
        elif action == 'save_stock':
            for entry in entries:
                new_stock_key = f'new_stock_{entry.id}'
                closing_key = f'closing_stock_{entry.id}'
                
                try:
                    new_val = int(request.form.get(new_stock_key, 0))
                except ValueError:
                    new_val = 0
                    
                entry.new_stock = new_val
                total_available = entry.opening_stock + new_val
                entry.total_available_stock = total_available
                
                # Optionally save valid closing_stock progress
                if closing_key in request.form and request.form[closing_key].strip() != '':
                    try:
                        closing_val = int(request.form[closing_key])
                        if 0 <= closing_val <= total_available:
                            entry.closing_stock = closing_val
                            entry.sold = total_available - closing_val
                    except ValueError:
                        pass
                        
                entry.is_completed = False
            db.session.commit()
            flash('Data saved successfully', 'success')
            return redirect(url_for('staff_dashboard'))
        else:
            entry_type_value = 'Estimated' if action == 'estimated' else 'Regular'
            for entry in entries:
                new_stock_key = f'new_stock_{entry.id}'
                closing_key = f'closing_stock_{entry.id}'
                
                try:
                    new_val = int(request.form.get(new_stock_key, 0))
                except ValueError:
                    new_val = 0
                
                entry.new_stock = new_val
                total_available = entry.opening_stock + new_val
                entry.total_available_stock = total_available
                
                if closing_key in request.form and request.form[closing_key].strip() != '':
                    try:
                        closing_val = int(request.form[closing_key])
                        if closing_val > total_available or closing_val < 0 or new_val < 0:
                            flash(f'Invalid stock data for {entry.product.brand_name}! Closing stock cannot exceed total available.', 'danger')
                            return redirect(url_for('staff_dashboard'))
                        
                        entry.closing_stock = closing_val
                        entry.sold = total_available - closing_val
                    except ValueError:
                        pass
                
                entry.entry_type = entry_type_value
                if action == 'estimated':
                    entry.is_completed = False
                else:
                    entry.is_completed = True
                    
            db.session.commit()
            flash(f'Data submitted successfully for {target_date.strftime("%Y-%m-%d")}', 'success')
            return redirect(url_for('staff_dashboard'))
    
    # Fetch entries (if not up_to_date)
    entries = []
    grouped_entries = {}
    if not is_up_to_date:
        entries = StockEntry.query.filter_by(date=target_date).join(Product).filter(Product.is_active == True).join(Category).order_by(Category.category_order).all()
        
        # Build raw dict structure: Category -> Brand -> [Entries]
        cat_brand_groups = {}
        for entry in entries:
            cat_name = entry.product.category.name
            brand_name = entry.product.brand_name
            
            if cat_name not in cat_brand_groups:
                cat_brand_groups[cat_name] = {}
            if brand_name not in cat_brand_groups[cat_name]:
                cat_brand_groups[cat_name][brand_name] = []
            
            cat_brand_groups[cat_name][brand_name].append(entry)
            
        # Re-map dynamically sorting internal brands & ML arrays
        grouped_entries = {}
        for cname, b_dict in cat_brand_groups.items():
            sorted_brands = []
            for bname, arr in b_dict.items():
                # Sort natively by ML size downward
                arr.sort(key=lambda x: getattr(x.product.ml_size, 'value', 0), reverse=True)
                # Max value sorting identifier for the brand block
                max_sp = max([x.selling_price_at_entry for x in arr]) if arr else 0
                sorted_brands.append({
                    'brand_name': bname,
                    'max_sp': max_sp,
                    'entries': arr
                })
            # Reorder the brand array strictly pointing down from `max_sp` computation, tiebreak by name
            sorted_brands.sort(key=lambda x: (-x['max_sp'], x['brand_name']))
            
            grouped_entries[cname] = []
            for brand_obj in sorted_brands:
                grouped_entries[cname].extend(brand_obj['entries'])
            
        
        # Display an alert if it is behind the working date significantly
        if target_date < working_date:
            flash(f'Action Required: You missed the entry for {target_date.strftime("%Y-%m-%d")}. You must complete it to proceed!', 'warning')
            
    return render_template('staff_dashboard.html', 
                           grouped_entries=grouped_entries,
                           target_date=target_date, 
                           working_date=working_date, 
                           is_up_to_date=is_up_to_date)

@app.route('/owner/download_report', methods=['POST'])
@login_required
@owner_required
def download_report():
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not start_date_str or not end_date_str:
        flash('Please select both Start and End dates.', 'danger')
        return redirect(url_for('owner_dashboard'))
        
    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format.', 'danger')
        return redirect(url_for('owner_dashboard'))
        
    entries_all = StockEntry.query.filter(StockEntry.date >= start_date, StockEntry.date <= end_date, StockEntry.is_completed == True).join(Product).join(Category).order_by(StockEntry.date, Category.category_order, Product.brand_name).all()
    
    if not entries_all:
        flash('No data available for the selected date range.', 'warning')
        return redirect(url_for('owner_dashboard'))
        
    wb = openpyxl.Workbook()
    # Remove default sheet initially and build dynamically
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(['Date', 'Total Sales', 'Total Profit'])
    
    grand_total_sold = 0
    grand_total_sales_amount = 0.0
    grand_total_profit = 0.0
    
    delta = end_date - start_date
    for i in range(delta.days + 1):
        target_d = start_date + timedelta(days=i)
        
        # Filter entries physically bounding inside python memory
        day_entries = [e for e in entries_all if e.date == target_d]
        if not day_entries:
            continue
            
        sheet_title = target_d.strftime('%d-%m-%Y')
        ws = wb.create_sheet(title=sheet_title)
        
        headers = ['Product', 'Category', 'ML', 'OB', 'New Stock', 'CB', 'Sold', 'Selling Price', 'Sale Amount', 'Profit']
        ws.append(headers)
        
        day_sold = 0
        day_sale_amount = 0.0
        day_profit = 0.0
        
        for entry in day_entries:
            sold = entry.sold
            selling_price = entry.selling_price_at_entry
            cost_price = entry.cost_price_at_entry
            sale_amount = sold * selling_price
            profit = sold * (selling_price - cost_price)
            
            row = [
                entry.product.brand_name,
                entry.product.category.name,
                f"{entry.product.ml_size.value} ml",
                entry.opening_stock,
                entry.new_stock,
                entry.closing_stock,
                sold,
                selling_price,
                sale_amount,
                profit
            ]
            ws.append(row)
            
            day_sold += sold
            day_sale_amount += sale_amount
            day_profit += profit
            
        ws.append([])
        ws.append(['TOTALS', '', '', '', '', '', day_sold, '', day_sale_amount, day_profit])
        
        # Append to summary log
        summary_ws.append([target_d.strftime('%Y-%m-%d'), day_sale_amount, day_profit])
        
        grand_total_sold += day_sold
        grand_total_sales_amount += day_sale_amount
        grand_total_profit += day_profit
        
    summary_ws.append([])
    summary_ws.append(['GRAND TOTAL', grand_total_sales_amount, grand_total_profit])
    
    # Push Summary to the front natively
    wb.move_sheet(summary_ws, offset=-len(wb.sheetnames))
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"sales_report_{start_date_str}_to_{end_date_str}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/owner/dashboard')
@login_required
@owner_required
def owner_dashboard():
    # Show stats for the LAST COMPLETED date, instead of today.
    # If the user wants to see "Today's" sales, it technically hasn't happened yet until closed.
    # Let's show stats for the last completed date as the default view.
    last_entry = StockEntry.query.filter_by(is_completed=True).order_by(StockEntry.date.desc()).first()
    
    report_date = last_entry.date if last_entry else date.today() - timedelta(days=1)
    
    entries_today = StockEntry.query.filter_by(date=report_date, is_completed=True).all()
    expenses_today = Expense.query.filter_by(date=report_date).all()
    
    total_sales = sum(e.sold for e in entries_today)
    total_revenue = sum(e.sold * e.selling_price_at_entry for e in entries_today)
    total_profit = sum(e.sold * (e.selling_price_at_entry - e.cost_price_at_entry) for e in entries_today)
    total_expenses = sum(exp.amount for exp in expenses_today)
    total_purchases = sum(e.new_stock * e.cost_price_at_entry for e in entries_today)
    purchased_bottles = sum(e.new_stock for e in entries_today)
    net_profit = total_profit - total_expenses
    
    # Audit trail: Show any dates that were estimated in the last 7 days
    seven_days_ago = date.today() - timedelta(days=7)
    estimated_entries = StockEntry.query.filter(
        StockEntry.is_completed == True,
        StockEntry.entry_type == 'Estimated',
        StockEntry.date >= seven_days_ago
    ).group_by(StockEntry.date).all()
    
    estimated_dates = [e.date.strftime("%Y-%m-%d") for e in estimated_entries]

    return render_template('owner_dashboard.html', 
        sales=total_sales, revenue=total_revenue,
        profit=total_profit, expenses=total_expenses, net_profit=net_profit, 
        purchases=total_purchases, purchased_bottles=purchased_bottles,
        report_date=report_date, estimated_dates=estimated_dates)

@app.route('/api/sales_chart_data')
@login_required
@owner_required
def sales_chart_data():
    last_entry = StockEntry.query.filter_by(is_completed=True).order_by(StockEntry.date.desc()).first()
    report_date = last_entry.date if last_entry else date.today() - timedelta(days=1)
    
    entries = StockEntry.query.filter_by(date=report_date, is_completed=True).join(Product).join(Category).all()
    cat_sales = {}
    for e in entries:
        cname = e.product.category.name
        cat_sales[cname] = cat_sales.get(cname, 0) + (e.sold * e.selling_price_at_entry)
        
    return jsonify({
        'labels': list(cat_sales.keys()),
        'data': list(cat_sales.values())
    })

@app.route('/owner/products', methods=['GET', 'POST'])
@login_required
@owner_required
def products():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            brand_name = request.form.get('brand_name')
            cat_id = int(request.form.get('category_id'))
            ml_id = int(request.form.get('ml_id'))
            cp = float(request.form.get('cost_price'))
            sp = float(request.form.get('selling_price'))
            
            # UNIQUE CONSTRAINT VALIDATION
            existing_product = Product.query.filter_by(brand_name=brand_name, category_id=cat_id, ml_id=ml_id).first()
            if existing_product:
                flash(f'Duplicate Blocked: {brand_name} internally mapped at that exact ML size.', 'danger')
                return redirect(url_for('products'))
            
            p = Product(brand_name=brand_name, category_id=cat_id, ml_id=ml_id, cost_price=cp, selling_price=sp)
            db.session.add(p)
            db.session.commit()
            flash('Product added successfully', 'success')
        elif action == 'edit':
            p_id = request.form.get('product_id')
            p = Product.query.get(p_id)
            if p:
                p.brand_name = request.form.get('brand_name')
                p.cost_price = float(request.form.get('cost_price'))
                p.selling_price = float(request.form.get('selling_price'))
                db.session.commit()
                flash('Product updated successfully', 'success')
        return redirect(url_for('products'))
        
    products_list = Product.query.join(Category).join(MLSize).order_by(Category.name, Product.brand_name, MLSize.value).all()
    categories = Category.query.all()
    return render_template('products.html', products=products_list, categories=categories)

@app.route('/api/toggle_product/<int:product_id>', methods=['POST'])
@login_required
@owner_required
def toggle_product(product_id):
    p = Product.query.get_or_404(product_id)
    # Block native physical deletions, toggle mapping correctly
    p.is_active = not p.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': p.is_active})

@app.route('/api/get_mls/<int:category_id>')
@login_required
@owner_required
def get_mls(category_id):
    mappings = CategoryMLMapping.query.filter_by(category_id=category_id).all()
    mls = [{'id': m.ml_id, 'value': MLSize.query.get(m.ml_id).value} for m in mappings]
    return jsonify(mls)

@app.route('/owner/expenses', methods=['GET', 'POST'])
@login_required
@owner_required
def expenses():
    # Expenses apply to working_date (yesterday) to match the stock logic
    working_date = date.today() - timedelta(days=1)
    if request.method == 'POST':
        amount = float(request.form.get('amount'))
        note = request.form.get('note')
        e = Expense(date=working_date, amount=amount, note=note)
        db.session.add(e)
        db.session.commit()
        flash('Expense added', 'success')
        return redirect(url_for('expenses'))
        
    expenses_working = Expense.query.filter_by(date=working_date).all()
    return render_template('expenses.html', expenses=expenses_working, working_date=working_date)

@app.route('/owner/override', methods=['GET', 'POST'])
@login_required
@owner_required
def override_mode():
    dates_query = db.session.query(StockEntry.date).filter_by(is_completed=True).distinct().order_by(StockEntry.date.desc()).all()
    available_dates = [d[0] for d in dates_query]

    selected_date_str = request.args.get('date') or request.form.get('date')
    selected_date = None
    entries = []
    
    if selected_date_str:
        try:
            selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            if request.method == 'POST':
                # Process override submission
                entries_to_update = StockEntry.query.filter_by(date=selected_date, is_completed=True).all()
                for entry in entries_to_update:
                    new_key = f"new_stock_{entry.id}"
                    closing_key = f"closing_stock_{entry.id}"
                    
                    if new_key in request.form and closing_key in request.form:
                        try:
                            n_val = int(request.form[new_key])
                            c_val = int(request.form[closing_key])
                            t_avail = entry.opening_stock + n_val
                            if c_val <= t_avail and c_val >= 0 and n_val >= 0:
                                entry.new_stock = n_val
                                entry.closing_stock = c_val
                                entry.total_available_stock = t_avail
                                entry.sold = t_avail - c_val
                        except ValueError:
                            continue
                db.session.commit()
                flash(f'Stock entries for {selected_date} have been overridden and forcefully updated.', 'success')
                return redirect(url_for('override_mode', date=selected_date_str))

            entries = StockEntry.query.filter_by(date=selected_date, is_completed=True).join(Product).join(Category).order_by(Category.category_order, Product.brand_name).all()
        except ValueError:
            flash('Invalid date selected.', 'danger')

    # Group entries identically to staff_dashboard for familiarity
    grouped_entries = {}
    if entries:
        cat_brand_groups = {}
        for entry in entries:
            cat_name = entry.product.category.name
            brand_name = entry.product.brand_name
            cat_brand_groups.setdefault(cat_name, {}).setdefault(brand_name, []).append(entry)
            
        for cname, b_dict in cat_brand_groups.items():
            sorted_brands = []
            for bname, arr in b_dict.items():
                arr.sort(key=lambda x: getattr(x.product.ml_size, 'value', 0), reverse=True)
                max_sp = max([x.selling_price_at_entry for x in arr]) if arr else 0
                sorted_brands.append({
                    'brand_name': bname,
                    'max_sp': max_sp,
                    'entries': arr
                })
            sorted_brands.sort(key=lambda x: (-x['max_sp'], x['brand_name']))
            
            grouped_entries[cname] = []
            for brand_obj in sorted_brands:
                grouped_entries[cname].extend(brand_obj['entries'])

    return render_template('override.html', 
                           available_dates=available_dates, 
                           selected_date=selected_date, 
                           grouped_entries=grouped_entries)

@app.route('/owner/reset_system', methods=['POST'])
@login_required
@owner_required
def reset_system():
    # Delete all Stock Entries and Expenses
    db.session.query(StockEntry).delete()
    db.session.query(Expense).delete()
    db.session.commit()
    flash('System reset successfully. All stock operations and audit logs erased.', 'success')
    return redirect(url_for('owner_dashboard'))

init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
